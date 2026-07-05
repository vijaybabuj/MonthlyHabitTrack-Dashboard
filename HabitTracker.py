import calendar
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TARGET_DATE = date.today().replace(day=1)
TARGET_MONTH_LABEL = TARGET_DATE.strftime("%B %Y")
MONTH_LENGTH = calendar.monthrange(TARGET_DATE.year, TARGET_DATE.month)[1]

# Create workbook and select active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "A4 Desk Tracker"

# Ensure grid lines are visible when printed and viewed
ws.views.sheetView[0].showGridLines = True

# Page setup for A4 Landscape
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins.left = 0.2
ws.page_margins.right = 0.2
ws.page_margins.top = 0.2
ws.page_margins.bottom = 0.2
ws.print_options.gridLines = True
ws.print_options.horizontalCentered = True
ws.freeze_panes = "C4"
ws.print_title_rows = "1:3"
ws.print_area = "A1:AG35"

# Define professional color palette
HEADER_BLUE = "1F4E78"     # Dark Steel Blue for Main Banner
MONTHLY_BLUE = "2F5597"    # Classic Navy for Monthly Goals
WEEKLY_GREEN = "385723"    # Forest Green for Weekly Goals
DAILY_ORANGE = "C65911"    # Rust Orange for Daily Goals
NOTES_GRAY = "595959"      # Charcoal Gray for Notes
SUBHEADER_BG = "F2F2F2"    # Very light gray for subheaders
WEEKEND_BG = "F2F4F7"      # Soft blue-gray for weekends
BORDER_COLOR = "D3D3D3"    # Light gray for grid borders

# Fonts
font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
font_section = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_sub = Font(name="Segoe UI", size=9, bold=True, color="595959")
font_body = Font(name="Segoe UI", size=9, color="000000")
font_body_bold = Font(name="Segoe UI", size=9, bold=True, color="000000")

# Fills
fill_title = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
fill_monthly = PatternFill(start_color=MONTHLY_BLUE, end_color=MONTHLY_BLUE, fill_type="solid")
fill_weekly = PatternFill(start_color=WEEKLY_GREEN, end_color=WEEKLY_GREEN, fill_type="solid")
fill_daily = PatternFill(start_color=DAILY_ORANGE, end_color=DAILY_ORANGE, fill_type="solid")
fill_notes = PatternFill(start_color=NOTES_GRAY, end_color=NOTES_GRAY, fill_type="solid")
fill_sub = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")
fill_weekend = PatternFill(start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid")

# Alignments
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Borders
thin_side = Side(border_style="thin", color=BORDER_COLOR)
border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

def style_range(ws, cell_range, font=None, fill=None, border=None, alignment=None):
    """Applies styles to a range of cells (handles merged ranges correctly)"""
    cells = ws[cell_range]
    if not isinstance(cells, tuple):
        cells = ((cells,),)

    for row in cells:
        for cell in row:
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment

# --- 1. MAIN TITLE BANNER (Row 1) ---
ws.merge_cells("A1:AG1")
ws["A1"] = f"MONTH: {TARGET_MONTH_LABEL}   |   THEME & MAIN FOCUS: Consistency, Energy & Focus"
style_range(ws, "A1:AG1", font=font_title, fill=fill_title, alignment=align_center)
ws.row_dimensions[1].height = 30

# Row 2 (Spacer)
ws.row_dimensions[2].height = 10

# --- 2. ZONE 3: DAILY HABIT & COMMITMENT TRACKER (Rows 3-15) ---
ws.merge_cells("A3:B3")
ws["A3"] = "DAILY HABITS & COMMITMENTS"
style_range(ws, "A3:B3", font=font_section, fill=fill_daily, alignment=align_left)

for day in range(1, 32):
    col_letter = get_column_letter(day + 2) # Starts at Col C (3)
    ws[f"{col_letter}3"] = day if day <= MONTH_LENGTH else ""
    style_range(ws, f"{col_letter}3", font=font_section, fill=fill_daily, alignment=align_center)

ws.row_dimensions[3].height = 22

daily_habits = [
    ("Sleep 8 Hours", "Everyday"),
    ("Meditate", "Mon-Fri"),
    ("Journaling / Reflection", "Everyday"),
    ("Morning Stretch", "Everyday"),
    ("Deep Work (2 Hours)", "Mon-Fri"),
    ("Learn (1 Hour)", "Mon-Fri"),
    ("Drink 2L Water", "Everyday"),
    ("Eat before 6 pm", "6 days/wk"),
    ("Random words & Questions", "Everyday"),
    ("Kids Edu-Review", "Mon-Fri"),
    ("Read 10+ Pages", "Everyday"),
    ("Limit Screen Time (<2h)", "Everyday"),
    ("", ""), # Blank row for custom habit
    ]
# Highlight weekends dynamically for the selected month.
weekends = {
    day
    for day in range(1, MONTH_LENGTH + 1)
    if date(TARGET_DATE.year, TARGET_DATE.month, day).weekday() >= 5
}

for r_idx, (habit, target) in enumerate(daily_habits, start=4):
    ws[f"A{r_idx}"] = habit
    ws[f"B{r_idx}"] = target
    style_range(ws, f"A{r_idx}", font=font_body_bold if habit else font_body, alignment=align_left)
    style_range(ws, f"B{r_idx}", font=font_sub, fill=fill_sub, alignment=align_center)
    
    # Set borders and weekend shading
    for day in range(1, 32):
        col_letter = get_column_letter(day + 2)
        cell = ws[f"{col_letter}{r_idx}"]
        cell.border = border_all_thin
        if day > MONTH_LENGTH:
            cell.fill = fill_sub
        elif day in weekends:
            cell.fill = fill_weekend
            
    ws.row_dimensions[r_idx].height = 20

# Apply borders to A and B columns for daily habits
for r in range(4, 4 + len(daily_habits)):
    ws[f"A{r}"].border = border_all_thin
    ws[f"B{r}"].border = border_all_thin

# Spacer rows below daily tracker
ws.row_dimensions[16].height = 10
ws.row_dimensions[17].height = 10

# --- 3. ZONE 1: MONTHLY MAJOR COMMITMENTS (Rows 18-24) ---
ws.merge_cells("A18:B18")
ws["A18"] = "MONTHLY MAJOR COMMITMENTS (Deadline Driven)"
style_range(ws, "A18:B18", font=font_section, fill=fill_monthly, alignment=align_center)

ws["A19"] = "Commitment / Goal"
ws["B19"] = "Due Date"
style_range(ws, "A19:A19", font=font_sub, fill=fill_sub, alignment=align_left)
style_range(ws, "B19:B19", font=font_sub, fill=fill_sub, alignment=align_center)

monthly_tasks = [
    ("", ""),
    ("Medical check-up", "Jul 31"),
    ("Read 1 book (Atomic Habits)", "May 31"),
    ("", ""),
    ("", "")
]

for idx, (task, due) in enumerate(monthly_tasks, start=20):
    ws[f"A{idx}"] = task
    ws[f"B{idx}"] = due
    style_range(ws, f"A{idx}:A{idx}", font=font_body, alignment=align_left)
    style_range(ws, f"B{idx}:B{idx}", font=font_body, alignment=align_center)

for r in range(18, 25):
    for c in range(1, 3):
        ws.cell(row=r, column=c).border = border_all_thin

# Spacer row below monthly tracker
ws.row_dimensions[25].height = 10

# --- 4. ZONE 2: WEEKLY ROUTINE TRACKER (Rows 26-32) ---
ws.merge_cells("A26:G26")
ws["A26"] = "WEEKLY ROUTINE TRACKER"
style_range(ws, "A26:G26", font=font_section, fill=fill_weekly, alignment=align_center)

ws["A27"] = "Weekly Task / Routine"
ws["B27"] = "Target"
ws["C27"] = "Wk 1"
ws["D27"] = "Wk 2"
ws["E27"] = "Wk 3"
ws["F27"] = "Wk 4"
ws["G27"] = "Wk 5"
style_range(ws, "A27:G27", font=font_sub, fill=fill_sub, alignment=align_center)
ws["A27"].alignment = align_left

weekly_tasks = [
    ("Run", "3x/wk"),
    ("Weekly Review & Planning", "1x/wk"),
    ("Review Budget & Expenses", "1x/wk"),
    ("FrogTech-planning", "2x/wk"),
    ("MTS Operations meeting & Prints", "1x/wk"),
]

for idx, (task, target) in enumerate(weekly_tasks, start=28):
    ws[f"A{idx}"] = task
    ws[f"B{idx}"] = target
    style_range(ws, f"A{idx}:B{idx}", font=font_body, alignment=align_left)
    ws[f"B{idx}"].alignment = align_center
    for col_idx in range(3, 8): # C to G
        style_range(ws, f"{get_column_letter(col_idx)}{idx}", font=font_body, alignment=align_center)

for r in range(26, 28 + len(weekly_tasks)):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border_all_thin

# --- 5. ZONE 4: NOTES & REFLECTIONS (Rows 18-35, Columns I-AG) ---
ws.merge_cells("I18:AG18")
ws["I18"] = "NOTES & REFLECTIONS"
style_range(ws, "I18:AG18", font=font_section, fill=fill_notes, alignment=align_center)

ws.merge_cells("I19:AG35")
ws["I19"] = "Use this space for thoughts, mid-month adjustments, or tracking wins!"
style_range(ws, "I19:AG35", font=Font(name="Segoe UI", size=9, italic=True, color="7F7F7F"), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True))

for r in range(18, 36):
    for c in range(9, 34):
        ws.cell(row=r, column=c).border = border_all_thin

# --- COLUMN WIDTH ADJUSTMENTS (Optimized for Landscape A4) ---
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 12
for day in range(1, 32):
    col_letter = get_column_letter(day + 2)
    ws.column_dimensions[col_letter].width = 3.5

ws.column_dimensions['H'].width = 3    # Spacer before notes
for col in ['C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col].width = 4.5

# Save Workbook
output_file = f"A4_Desk_Habit_Tracker_{TARGET_DATE.strftime('%Y_%m')}.xlsx"
wb.save(output_file)
print("File successfully generated!")
